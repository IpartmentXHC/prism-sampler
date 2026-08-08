from __future__ import annotations

import hashlib
import json
import math
import re
import shutil
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any, Iterable

import duckdb


SCHEMA_VERSION = "affinitygraph.placement.v2"
CSV_VIEWS = {
    "solve_windows": "windows.csv",
    "threads": "threads.csv",
    "groups": "thread_groups.csv",
    "edges": "edges.csv",
    "plans": "plans.csv",
    "family_metrics": "family_metrics.csv",
    "domains": "domains.csv",
    "planned_masks": "planned_masks.csv",
    "mask_actions": "mask_actions.csv",
    "cpu_loads": "cpu_distribution.csv",
    "action_batches": "action_batches.csv",
    "bpf_health": "bpf_health.csv",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parse_cpu_mask(value: str) -> list[int]:
    cpus: set[int] = set()
    for item in str(value).split(","):
        if not item:
            continue
        if "-" in item:
            start, end = (int(part) for part in item.split("-", 1))
            cpus.update(range(start, end + 1))
        else:
            cpus.add(int(item))
    return sorted(cpus)


def _runtime_logs(run: Path) -> list[Path]:
    if run.is_file():
        return [run]
    logs = sorted(path for path in run.rglob("runtime.jsonl") if path.is_file())
    if not logs:
        logs = sorted(path for path in run.rglob("*.jsonl") if path.is_file())
    if not logs:
        raise ValueError(f"no JSONL logs found under {run}")
    return logs


def _read_events(paths: Iterable[Path]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    for path in paths:
        with path.open(encoding="utf-8") as stream:
            for line_number, line in enumerate(stream, 1):
                if not line.strip():
                    continue
                try:
                    event = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"{path}:{line_number}: invalid JSON: {exc}") from exc
                if not isinstance(event, dict) or "type" not in event:
                    raise ValueError(f"{path}:{line_number}: event object/type required")
                event["_source"] = str(path)
                events.append(event)
    events.sort(key=lambda row: (int(row.get("timestamp_ns", 0)), row["_source"]))
    return events


def _windowed_events(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    active: str | None = None
    for event in events:
        if event["type"] == "solve_window_begin":
            active = str(event["window_id"])
        elif "window_id" not in event and active and event["type"] in {
            "thread_window", "relation_edge", "plan", "action", "bpf_health"
        }:
            event["window_id"] = active
        if event["type"] == "solve_window_end":
            active = None
    return events


def _frame_legacy_plan_windows(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    root_pid = next((int(row["root_pid"]) for row in events
                     if row["type"] == "runtime_start" and row.get("root_pid")), -1)
    synthetic: list[dict[str, Any]] = []
    for index, plan in enumerate(events):
        if plan["type"] != "plan" or plan.get("window_id"):
            continue
        cursor = index - 1
        edge_rows: list[dict[str, Any]] = []
        while cursor >= 0 and events[cursor]["type"] == "relation_edge":
            edge_rows.append(events[cursor])
            cursor -= 1
        thread_rows: list[dict[str, Any]] = []
        while cursor >= 0 and events[cursor]["type"] == "thread_window":
            thread_rows.append(events[cursor])
            cursor -= 1
        if not thread_rows:
            continue
        window_id = f"legacy_unframed-{plan['timestamp_ns']}"
        for row in thread_rows:
            row["window_id"] = window_id
            row.setdefault("tgid", root_pid)
        for row in edge_rows:
            row["window_id"] = window_id
        plan["window_id"] = window_id
        plan["legacy_reconstructed"] = True
        plan.setdefault("strategy_id", "legacy-v1")
        begin_ns = min(int(row["timestamp_ns"]) for row in thread_rows + edge_rows)
        synthetic.extend([
            {"timestamp_ns": begin_ns, "type": "solve_window_begin",
             "window_id": window_id, "sequence": len(synthetic) // 2 + 1,
             "mode": "plan", "legacy_reconstructed": True,
             "_source": plan["_source"]},
            {"timestamp_ns": int(plan["timestamp_ns"]), "type": "solve_window_end",
             "window_id": window_id, "complete": True,
             "outcome": "legacy_reconstructed", "_source": plan["_source"]},
        ])
    return sorted(events + synthetic,
                  key=lambda row: (int(row.get("timestamp_ns", 0)), row["_source"], row["type"]))


def _infer_topology(run: Path, timestamp_ns: int) -> list[dict[str, Any]]:
    if not run.is_dir():
        return []
    events: list[dict[str, Any]] = []
    lscpu = next(iter(sorted(run.rglob("lscpu-e.txt"))), None)
    if lscpu:
        for line in lscpu.read_text().splitlines()[1:]:
            fields = line.split()
            if len(fields) >= 5 and fields[0].isdigit() and fields[1].lstrip("-").isdigit():
                events.append({"timestamp_ns": timestamp_ns, "type": "topology_cpu",
                               "cpu": int(fields[0]), "node": int(fields[1]),
                               "online": fields[4].lower() == "yes", "in_envelope": True,
                               "_source": str(lscpu)})
    numactl = next(iter(sorted(run.rglob("numactl-hardware.txt"))), None)
    if numactl:
        lines = numactl.read_text().splitlines()
        for line in lines:
            match = re.match(r"\s*(\d+):\s+((?:\d+\s*)+)$", line)
            if not match:
                continue
            from_node = int(match.group(1))
            for to_node, distance in enumerate(int(value) for value in match.group(2).split()):
                events.append({"timestamp_ns": timestamp_ns, "type": "topology_edge",
                               "from_node": from_node, "to_node": to_node,
                               "numa_distance": distance, "_source": str(numactl)})
    return events


def _legacy_unframed_canaries(events: list[dict[str, Any]], envelope_cpu_count: int) -> list[dict[str, Any]]:
    canaries: list[dict[str, Any]] = []
    for row in events:
        if (row["type"] not in {"plan", "action"} or
                (row.get("window_id") and not row.get("legacy_reconstructed")) or
                not row.get("assignments")):
            continue
        counts: dict[int, int] = {}
        for cpu in row["assignments"].values():
            counts[int(cpu)] = counts.get(int(cpu), 0) + 1
        thread_count = sum(counts.values())
        dynamic_cap = math.ceil(thread_count / max(envelope_cpu_count, 1)) + 2
        maximum = max(counts.values(), default=0)
        canaries.append({
            "classification": "legacy_unframed", "timestamp_ns": row.get("timestamp_ns"),
            "thread_count": thread_count, "max_threads_on_one_cpu": maximum,
            "dynamic_slot_cap": dynamic_cap, "pathological_concentration": maximum > dynamic_cap,
        })
    return canaries


def _create_schema(connection: duckdb.DuckDBPyConnection) -> None:
    connection.execute("""
        CREATE TABLE manifest(key VARCHAR PRIMARY KEY, value VARCHAR NOT NULL);
        CREATE TABLE solve_windows(
            window_id VARCHAR PRIMARY KEY, begin_ns UBIGINT NOT NULL, end_ns UBIGINT,
            sequence BIGINT, mode VARCHAR, complete BOOLEAN NOT NULL, outcome VARCHAR,
            source_log VARCHAR NOT NULL
        );
        CREATE TABLE threads(
            window_id VARCHAR NOT NULL REFERENCES solve_windows(window_id),
            tgid BIGINT NOT NULL, tid BIGINT NOT NULL, starttime UBIGINT NOT NULL,
            comm VARCHAR, parent_tid BIGINT, start_routine UBIGINT, start_symbol VARCHAR,
            group_name VARCHAR, demand DOUBLE, confidence DOUBLE, current_cpu INTEGER,
            allowed_cpus INTEGER[], state VARCHAR, sample_count BIGINT,
            runtime_delta_ns UBIGINT, runqueue_delta_ns UBIGINT,
            voluntary_switches_delta UBIGINT, involuntary_switches_delta UBIGINT,
            PRIMARY KEY(window_id, tgid, tid, starttime)
        );
        CREATE TABLE groups(
            window_id VARCHAR NOT NULL REFERENCES solve_windows(window_id),
            group_name VARCHAR NOT NULL, thread_count BIGINT NOT NULL,
            total_demand DOUBLE NOT NULL, mean_confidence DOUBLE,
            PRIMARY KEY(window_id, group_name)
        );
        CREATE TABLE edges(
            window_id VARCHAR NOT NULL REFERENCES solve_windows(window_id),
            from_tid BIGINT NOT NULL, from_starttime UBIGINT NOT NULL,
            to_tid BIGINT NOT NULL, to_starttime UBIGINT NOT NULL,
            activity DOUBLE, sync DOUBLE, share DOUBLE, stability DOUBLE, score DOUBLE,
            handoff_rate DOUBLE, shared_vfs_seconds DOUBLE, active_overlap DOUBLE,
            observation_count BIGINT, coverage DOUBLE, cv DOUBLE,
            PRIMARY KEY(window_id, from_tid, from_starttime, to_tid, to_starttime)
        );
        CREATE TABLE topology_cpus(
            cpu INTEGER PRIMARY KEY, node INTEGER NOT NULL, online BOOLEAN NOT NULL,
            in_envelope BOOLEAN NOT NULL
        );
        CREATE TABLE topology_edges(
            from_node INTEGER, to_node INTEGER, numa_distance DOUBLE,
            handoff_mean_ns DOUBLE, handoff_p95_ns DOUBLE,
            memory_load_mean_ns DOUBLE, memory_load_cv DOUBLE,
            stream_2t_triad_mbps DOUBLE, stream_32t_triad_mbps DOUBLE,
            PRIMARY KEY(from_node, to_node)
        );
        CREATE TABLE plans(
            window_id VARCHAR NOT NULL REFERENCES solve_windows(window_id),
            strategy_id VARCHAR NOT NULL, thread_count BIGINT, confirmation_threads BIGINT,
            overload DOUBLE, relation_cost DOUBLE, migration_cost DOUBLE,
            confirmation BIGINT, PRIMARY KEY(window_id, strategy_id)
        );
        CREATE TABLE assignments(
            window_id VARCHAR NOT NULL, strategy_id VARCHAR NOT NULL,
            tgid BIGINT NOT NULL, tid BIGINT NOT NULL, starttime UBIGINT NOT NULL,
            cpu INTEGER NOT NULL,
            PRIMARY KEY(window_id, strategy_id, tgid, tid, starttime)
        );
        CREATE TABLE family_metrics(
            window_id VARCHAR NOT NULL, strategy_id VARCHAR NOT NULL,
            family_name VARCHAR NOT NULL, thread_count BIGINT, demand DOUBLE,
            internal_relation DOUBLE, external_relation DOUBLE,
            self_containment DOUBLE, relative_internal DOUBLE,
            confirmation BIGINT, anchor BOOLEAN,
            PRIMARY KEY(window_id, strategy_id, family_name)
        );
        CREATE TABLE domains(
            window_id VARCHAR NOT NULL, strategy_id VARCHAR NOT NULL,
            domain_id VARCHAR NOT NULL, families VARCHAR[], thread_count BIGINT,
            target_nodes INTEGER[], target_mask INTEGER[], demand DOUBLE,
            confirmation BIGINT, valid BOOLEAN, invalid_reason VARCHAR,
            PRIMARY KEY(window_id, strategy_id, domain_id)
        );
        CREATE TABLE planned_masks(
            window_id VARCHAR NOT NULL, strategy_id VARCHAR NOT NULL,
            tgid BIGINT NOT NULL, tid BIGINT NOT NULL, starttime UBIGINT NOT NULL,
            cpus INTEGER[] NOT NULL,
            PRIMARY KEY(window_id, strategy_id, tgid, tid, starttime)
        );
        CREATE TABLE mask_actions(
            timestamp_ns UBIGINT NOT NULL, window_id VARCHAR, tid BIGINT NOT NULL,
            from_mask INTEGER[], target_mask INTEGER[], target_nodes INTEGER[],
            forced_migration BOOLEAN
        );
        CREATE TABLE cpu_loads(
            window_id VARCHAR NOT NULL, strategy_id VARCHAR NOT NULL, cpu INTEGER NOT NULL,
            thread_count BIGINT NOT NULL, demand DOUBLE NOT NULL,
            PRIMARY KEY(window_id, strategy_id, cpu)
        );
        CREATE TABLE action_batches(
            timestamp_ns UBIGINT NOT NULL, window_id VARCHAR, kind VARCHAR NOT NULL,
            success BOOLEAN, requested BIGINT, committed BIGINT, vanished BIGINT,
            rolled_back BIGINT, error BIGINT, restore_requested BIGINT,
            restore_restored BIGINT, restore_vanished BIGINT, restore_failed BIGINT
        );
        CREATE TABLE bpf_health(
            timestamp_ns UBIGINT NOT NULL, window_id VARCHAR, valid BOOLEAN, error BIGINT,
            emitted UBIGINT, dropped UBIGINT, loss_ratio DOUBLE, window_ready BOOLEAN,
            final BOOLEAN
        );
    """)


def export_placement(run: Path, output: Path) -> dict[str, Any]:
    run = run.resolve()
    output = output.resolve()
    logs = _runtime_logs(run)
    events = _read_events(logs)
    events = _frame_legacy_plan_windows(events)
    if not any(row["type"] == "topology_cpu" for row in events):
        events.extend(_infer_topology(run, int(events[0].get("timestamp_ns", 0))))
    events = _windowed_events(events)
    output.mkdir(parents=True, exist_ok=True)
    database = output / "placement.duckdb"
    if database.exists():
        database.unlink()
    connection = duckdb.connect(str(database))
    _create_schema(connection)
    connection.execute("BEGIN TRANSACTION")

    begins = {str(row["window_id"]): row for row in events if row["type"] == "solve_window_begin"}
    ends = {str(row["window_id"]): row for row in events if row["type"] == "solve_window_end"}
    for window_id, begin in begins.items():
        end = ends.get(window_id)
        connection.execute(
            "INSERT INTO solve_windows VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            [window_id, int(begin["timestamp_ns"]), int(end["timestamp_ns"]) if end else None,
             begin.get("sequence"), begin.get("mode"), bool(end and end.get("complete")),
             end.get("outcome") if end else None, begin["_source"]],
        )

    identities: dict[tuple[str, int], tuple[int, int]] = {}
    thread_events = [row for row in events if row["type"] == "thread_window" and row.get("window_id") in begins]
    for row in thread_events:
        window_id = str(row["window_id"])
        identity = (int(row.get("tgid", -1)), int(row.get("starttime", 0)))
        identities[(window_id, int(row["tid"]))] = identity
        connection.execute("""
            INSERT INTO threads VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [window_id, identity[0], int(row["tid"]), identity[1], row.get("comm"),
              row.get("parent_tid"), row.get("start_routine"), row.get("start_symbol"),
              row.get("group", "unknown"), row.get("demand", 0.0), row.get("confidence", 0.0),
              row.get("current_cpu", -1), row.get("allowed_cpus"), row.get("state"),
              row.get("sample_count"), row.get("runtime_delta_ns"),
              row.get("runqueue_delta_ns"), row.get("voluntary_switches_delta"),
              row.get("involuntary_switches_delta")])
    connection.execute("""
        INSERT INTO groups
        SELECT window_id, group_name, count(*), sum(demand), avg(confidence)
        FROM threads GROUP BY window_id, group_name
    """)

    for row in (item for item in events if item["type"] == "relation_edge" and item.get("window_id") in begins):
        window_id = str(row["window_id"])
        from_identity = identities.get((window_id, int(row["from_tid"])))
        to_identity = identities.get((window_id, int(row["to_tid"])))
        if not from_identity or not to_identity:
            continue
        connection.execute("INSERT INTO edges VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [window_id, int(row["from_tid"]), from_identity[1], int(row["to_tid"]), to_identity[1],
             row.get("activity"), row.get("sync"), row.get("share"), row.get("stability"),
             row.get("score"), row.get("handoff_rate"), row.get("shared_vfs_seconds"),
             row.get("active_overlap"), row.get("observation_count"), row.get("coverage"), row.get("cv")])

    for row in (item for item in events if item["type"] == "topology_cpu"):
        connection.execute("INSERT OR REPLACE INTO topology_cpus VALUES (?, ?, ?, ?)",
            [row["cpu"], row["node"], row.get("online", True), row.get("in_envelope", True)])
    for row in (item for item in events if item["type"] == "topology_edge"):
        connection.execute("INSERT OR REPLACE INTO topology_edges VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [row["from_node"], row["to_node"], row.get("numa_distance"),
             row.get("handoff_mean_ns"), row.get("handoff_p95_ns"),
             row.get("memory_load_mean_ns"), row.get("memory_load_cv"),
             row.get("stream_2t_triad_mbps"), row.get("stream_32t_triad_mbps")])

    for row in (item for item in events if item["type"] == "plan" and item.get("window_id") in begins):
        window_id = str(row["window_id"])
        strategy = str(row.get("strategy_id", "legacy-v1"))
        connection.execute("INSERT INTO plans VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            [window_id, strategy, row.get("threads"), row.get("confirmation_threads"),
             row.get("overload"), row.get("relation_cost"), row.get("migration_cost"), row.get("confirmation")])
        for tid_text, cpu in row.get("assignments", {}).items():
            tid = int(tid_text)
            identity = identities.get((window_id, tid))
            if not identity:
                continue
            connection.execute("INSERT INTO assignments VALUES (?, ?, ?, ?, ?, ?)",
                [window_id, strategy, identity[0], tid, identity[1], int(cpu)])
        for family in row.get("family_metrics", []):
            connection.execute("INSERT INTO family_metrics VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", [
                window_id, strategy, family.get("name"), family.get("thread_count"),
                family.get("demand"), family.get("internal_relation"),
                family.get("external_relation"), family.get("self_containment"),
                family.get("relative_internal"), family.get("confirmation"),
                family.get("anchor"),
            ])
        for domain in row.get("domains", []):
            connection.execute("INSERT INTO domains VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", [
                window_id, strategy, domain.get("id"), domain.get("families", []),
                domain.get("thread_count"), domain.get("target_nodes", []),
                _parse_cpu_mask(domain.get("target_mask", "")), domain.get("demand"),
                domain.get("confirmation"), domain.get("valid"),
                domain.get("invalid_reason"),
            ])
        for tid_text, mask in row.get("planned_masks", {}).items():
            tid = int(tid_text)
            identity = identities.get((window_id, tid))
            if identity:
                connection.execute("INSERT INTO planned_masks VALUES (?, ?, ?, ?, ?, ?)", [
                    window_id, strategy, identity[0], tid, identity[1],
                    _parse_cpu_mask(mask),
                ])
    connection.execute("""
        INSERT INTO cpu_loads
        SELECT a.window_id, a.strategy_id, a.cpu, count(*), sum(t.demand)
        FROM assignments a JOIN threads t USING(window_id, tgid, tid, starttime)
        GROUP BY a.window_id, a.strategy_id, a.cpu
    """)

    for row in events:
        if row["type"] in {"action", "pause", "runtime_stop"}:
            connection.execute("INSERT INTO action_batches VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [row.get("timestamp_ns", 0), row.get("window_id"), row["type"], row.get("success"),
                 row.get("requested"), row.get("committed"), row.get("vanished"), row.get("rolled_back"),
                 row.get("error"), row.get("restore_requested"), row.get("restore_restored"),
                 row.get("restore_vanished"), row.get("restore_failed")])
            for action in row.get("actions", []):
                if "target_mask" not in action:
                    continue
                connection.execute("INSERT INTO mask_actions VALUES (?, ?, ?, ?, ?, ?, ?)", [
                    row.get("timestamp_ns", 0), row.get("window_id"), action["tid"],
                    _parse_cpu_mask(action.get("from_mask", "")),
                    _parse_cpu_mask(action.get("target_mask", "")),
                    action.get("target_nodes", []), action.get("forced_migration", False),
                ])
        elif row["type"] == "bpf_health":
            connection.execute("INSERT INTO bpf_health VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [row.get("timestamp_ns", 0), row.get("window_id"), row.get("valid"), row.get("error"),
                 row.get("emitted"), row.get("dropped"), row.get("loss_ratio"),
                 row.get("window_ready"), row.get("final")])

    hash_candidates = set(logs)
    if run.is_dir():
        for pattern in ("*.toml", "*manifest*.json", "*.bpf.o"):
            hash_candidates.update(path for path in run.rglob(pattern) if path.is_file())
    envelope_cpu_count = connection.execute(
        "SELECT count(*) FROM topology_cpus WHERE in_envelope"
    ).fetchone()[0]
    manifest = {
        "schema": SCHEMA_VERSION,
        "exported_at_ns": str(time.time_ns()),
        "run": str(run),
        "source_hashes": json.dumps({str(path.relative_to(run) if run.is_dir() else path.name): _sha256(path)
                                      for path in sorted(hash_candidates)}, sort_keys=True),
        "legacy_unframed_canaries": json.dumps(
            _legacy_unframed_canaries(events, envelope_cpu_count), sort_keys=True
        ),
    }
    connection.executemany("INSERT INTO manifest VALUES (?, ?)", manifest.items())
    connection.execute("COMMIT")

    snapshot_dir = output / "snapshots"
    snapshot_dir.mkdir(exist_ok=True)
    complete_windows = connection.execute(
        "SELECT window_id, begin_ns FROM solve_windows WHERE complete ORDER BY begin_ns"
    ).fetchall()
    cpu_rows = connection.execute(
        "SELECT cpu, node, online FROM topology_cpus WHERE in_envelope ORDER BY cpu"
    ).fetchall()
    distance_rows = connection.execute(
        "SELECT from_node, to_node, numa_distance, handoff_mean_ns FROM topology_edges ORDER BY 1, 2"
    ).fetchall()
    sequence_frames = []
    for window_id, begin_ns in complete_windows:
        thread_rows = connection.execute("""
            SELECT tgid, tid, starttime, group_name, demand, confidence, current_cpu
            FROM threads WHERE window_id = ? ORDER BY tid
        """, [window_id]).fetchall()
        edge_rows = connection.execute("""
            SELECT from_tid, to_tid, activity, sync, share, stability, score,
                   handoff_rate, shared_vfs_seconds, active_overlap, observation_count, coverage, cv
            FROM edges WHERE window_id = ? ORDER BY from_tid, to_tid
        """, [window_id]).fetchall()
        snapshot = {
            "schema": "affinitygraph.snapshot.v1", "window_id": window_id,
            "topology": {
                "cpus": [{"id": cpu, "node": node, "online": online} for cpu, node, online in cpu_rows],
                "node_distances": [{"from_node": row[0], "to_node": row[1], "distance": row[2]}
                                   for row in distance_rows],
                "cpu_latencies": [
                    {"from_cpu": from_cpu, "to_cpu": to_cpu,
                     "latency": 0.0 if from_cpu == to_cpu else handoff}
                    for from_cpu, from_node, _ in cpu_rows
                    for to_cpu, to_node, _ in cpu_rows
                    for edge_from, edge_to, _, handoff in distance_rows
                    if from_node == edge_from and to_node == edge_to and handoff is not None
                ],
            },
            "threads": [{"tgid": row[0], "tid": row[1], "starttime": row[2], "group": row[3],
                         "demand": row[4], "confidence": row[5], "current_cpu": row[6]}
                        for row in thread_rows],
            "edges": [dict(zip(("from_tid", "to_tid", "activity", "sync", "share", "stability", "score",
                                "handoff_rate", "shared_vfs_seconds", "active_overlap", "observation_count",
                                "coverage", "cv"), row)) for row in edge_rows],
        }
        snapshot_name = f"{window_id}.json"
        (snapshot_dir / snapshot_name).write_text(
            json.dumps(snapshot, indent=2, sort_keys=True) + "\n"
        )
        sequence_frames.append({
            "timestamp_ns": int(begin_ns),
            "snapshot": f"snapshots/{snapshot_name}",
        })
    sequence = {
        "schema": "affinitygraph.replay-sequence.v1",
        "frames": sequence_frames,
    }
    (output / "sequence.json").write_text(
        json.dumps(sequence, indent=2, sort_keys=True) + "\n"
    )

    for table, filename in CSV_VIEWS.items():
        target = output / filename
        connection.execute(f"COPY (SELECT * FROM {table}) TO ? (HEADER, DELIMITER ',')", [str(target)])
    connection.execute("""
        COPY (
            SELECT 'cpu' AS record_type, cpu, node, online, in_envelope,
                   NULL::INTEGER AS from_node, NULL::INTEGER AS to_node,
                   NULL::DOUBLE AS numa_distance, NULL::DOUBLE AS handoff_mean_ns,
                   NULL::DOUBLE AS handoff_p95_ns, NULL::DOUBLE AS memory_load_mean_ns,
                   NULL::DOUBLE AS memory_load_cv, NULL::DOUBLE AS stream_2t_triad_mbps,
                   NULL::DOUBLE AS stream_32t_triad_mbps
            FROM topology_cpus
            UNION ALL
            SELECT 'node_edge', NULL, NULL, NULL, NULL, from_node, to_node,
                   numa_distance, handoff_mean_ns, handoff_p95_ns, memory_load_mean_ns,
                   memory_load_cv, stream_2t_triad_mbps, stream_32t_triad_mbps
            FROM topology_edges
        ) TO ? (HEADER, DELIMITER ',')
    """, [str(output / "topology.csv")])
    counts = {table: connection.execute(f"SELECT count(*) FROM {table}").fetchone()[0]
              for table in CSV_VIEWS}
    counts["topology_cpus"] = connection.execute("SELECT count(*) FROM topology_cpus").fetchone()[0]
    counts["topology_edges"] = connection.execute("SELECT count(*) FROM topology_edges").fetchone()[0]
    connection.close()
    return {"schema": SCHEMA_VERSION, "database": str(database),
            "sequence": str(output / "sequence.json"), "snapshots": len(complete_windows),
            "counts": counts, "source_hashes": json.loads(manifest["source_hashes"])}


def _find_replay() -> Path:
    executable = shutil.which("affinity-replay")
    if executable:
        return Path(executable)
    local = Path(__file__).resolve().parents[2] / "affinitygraph" / "build" / "affinity-replay"
    if local.is_file():
        return local
    raise FileNotFoundError("affinity-replay not found in PATH or sibling affinitygraph/build")


def _candidate_files(path: Path) -> list[Path]:
    return sorted(path.glob("*.toml")) if path.is_dir() else [path]


def _dominates(left: dict[str, float], right: dict[str, float], metrics: list[str]) -> bool:
    return all(left[name] <= right[name] for name in metrics) and any(
        left[name] < right[name] for name in metrics
    )


def compare_placement(dataset: Path, candidates: Path, output: Path) -> dict[str, Any]:
    dataset = dataset.resolve()
    root = dataset.parent
    connection = duckdb.connect(str(dataset), read_only=True)
    windows = [row[0] for row in connection.execute(
        "SELECT window_id FROM solve_windows WHERE complete ORDER BY begin_ns"
    ).fetchall()]
    connection.close()
    if not windows:
        raise ValueError("dataset has no complete solve windows")
    strategies = _candidate_files(candidates.resolve())
    if not strategies:
        raise ValueError("no candidate TOML files found")
    replay = _find_replay()
    rows: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="affinitygraph-compare-") as temporary:
        temporary_path = Path(temporary)
        for strategy in strategies:
            for window_id in windows:
                result_path = temporary_path / f"{strategy.stem}-{window_id}.json"
                timings: list[float] = []
                result: dict[str, Any] | None = None
                for _ in range(5):
                    started = time.perf_counter()
                    subprocess.run([str(replay), "--snapshot", str(root / "snapshots" / f"{window_id}.json"),
                                    "--strategy", str(strategy), "--output", str(result_path)],
                                   check=True, capture_output=True, text=True)
                    timings.append((time.perf_counter() - started) * 1000)
                    current = json.loads(result_path.read_text())
                    if result and current["assignments"] != result["assignments"]:
                        current["gates"]["deterministic"] = False
                        current["gates"]["passed"] = False
                    result = current
                assert result is not None
                result["metrics"]["wall_solve_p95_ms"] = sorted(timings)[-1]
                result["gates"]["solve_p95_below_1s"] = result["metrics"]["wall_solve_p95_ms"] < 1000
                result["gates"]["passed"] = result["gates"]["passed"] and result["gates"]["solve_p95_below_1s"]
                rows.append(result)
    by_strategy: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_strategy.setdefault(row["strategy_id"], []).append(row)
    for strategy_rows in by_strategy.values():
        maximum_churn = 0.0
        for previous, current in zip(strategy_rows, strategy_rows[1:]):
            before = {(row["tgid"], row["tid"], row["starttime"]): row["cpu"]
                      for row in previous["assignments"]}
            after = {(row["tgid"], row["tid"], row["starttime"]): row["cpu"]
                     for row in current["assignments"]}
            shared = before.keys() & after.keys()
            if shared:
                maximum_churn = max(maximum_churn,
                                    sum(before[key] != after[key] for key in shared) / len(shared))
        for row in strategy_rows:
            row["metrics"]["continuous_window_churn"] = maximum_churn
    metric_names = ["max_threads_per_cpu", "p95_threads_per_cpu", "thread_count_cv", "thread_count_gini",
                    "max_cpu_demand", "relation_weighted_latency", "same_cpu_edge_weight",
                    "active_migrations", "inactive_migrations", "migration_cost", "migration_distance",
                    "continuous_window_churn", "group_nonlocality", "mean_group_cpu_spread",
                    "wall_solve_p95_ms"]
    summaries: list[dict[str, Any]] = []
    for strategy_id, strategy_rows in sorted(by_strategy.items()):
        metrics = {name: max(float(row["metrics"][name]) for row in strategy_rows) for name in metric_names}
        summaries.append({"strategy_id": strategy_id,
                          "passed": all(row["gates"]["passed"] for row in strategy_rows),
                          "windows": len(strategy_rows), "metrics": metrics,
                          "failed_gates": sorted({name for row in strategy_rows for name, passed in row["gates"].items()
                                                  if isinstance(passed, bool) and not passed})})
    eligible = [row for row in summaries if row["passed"]]
    pareto = [row["strategy_id"] for row in eligible if not any(
        other is not row and _dominates(other["metrics"], row["metrics"], metric_names)
        for other in eligible
    )]
    legacy = next((row for row in summaries if row["strategy_id"] == "legacy-v1"), None)
    for row in summaries:
        row["relative_to_legacy"] = ({name: row["metrics"][name] - legacy["metrics"][name]
                                      for name in metric_names} if legacy else None)
    report = {"schema": "affinitygraph.comparison.v1", "dataset": str(dataset),
              "hard_gate_then_pareto": True, "pareto_frontier": pareto,
              "candidates": summaries, "window_results": rows}
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n")
    return report


def export_placement_excel(dataset: Path, output: Path) -> dict[str, int]:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    connection = duckdb.connect(str(dataset.resolve()), read_only=True)
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    guide = workbook.create_sheet("Guide")
    guide_rows = [
        ("AffinityGraph canonical-v11 离线分析表", "这不是运行时数据库；它由旧 JSONL 事后导出。"),
        ("推荐阅读顺序", "Overview -> CPU Distribution -> Groups -> Threads -> Edges -> Plans"),
        ("window_id", "一次离线恢复的 10 秒 solve 周期；outcome=legacy_reconstructed 表示旧日志重建。"),
        ("demand", "线程的 CPU 运行+排队压力，约 0~1；不是 CPU 利用率百分比。"),
        ("confidence", "窗口覆盖置信度。优先看 >=0.8 的线程。"),
        ("group_name", "由线程名/启动入口/父线程归一化得到的线程组。"),
        ("score", "线程关系综合权重；应结合 activity/sync/share/stability 看，不能单独当因果。"),
        ("Top Edges", "Excel 每个窗口仅保留 score 最高的 200 条边；完整边集在 DuckDB/edges.csv。"),
        ("CPU thread_count", "先找单核线程数异常高的窗口；721/128 的参考上限是 8。"),
        ("注意", "旧格式没有 comm、allowed mask、sample delta 和新增关系证据，这些单元格保持空白。"),
        ("性能标签", "canonical-v11 主机 busy，吞吐结果不能用作算法优劣标签。"),
    ]
    for left, right in guide_rows:
        guide.append([left, right])
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 100

    queries = {
        "Overview": """
            WITH current_counts AS (
                SELECT window_id, current_cpu AS cpu, count(*) AS n, sum(demand) AS demand
                FROM threads GROUP BY window_id, current_cpu
            ), plan_counts AS (
                SELECT window_id, strategy_id, cpu, count(*) AS n
                FROM assignments GROUP BY window_id, strategy_id, cpu
            )
            SELECT w.window_id, w.outcome,
                   (SELECT count(*) FROM threads t WHERE t.window_id=w.window_id) AS threads,
                   (SELECT count(*) FROM threads t WHERE t.window_id=w.window_id AND confidence >= 0.8) AS eligible_threads,
                   (SELECT sum(demand) FROM threads t WHERE t.window_id=w.window_id) AS total_demand,
                   (SELECT max(n) FROM current_counts c WHERE c.window_id=w.window_id) AS max_threads_on_current_cpu,
                   (SELECT max(demand) FROM current_counts c WHERE c.window_id=w.window_id) AS max_current_cpu_demand,
                   (SELECT count(*) FROM edges e WHERE e.window_id=w.window_id) AS edge_count,
                   (SELECT sum(score) FROM edges e WHERE e.window_id=w.window_id) AS edge_weight,
                   (SELECT max(n) FROM plan_counts p WHERE p.window_id=w.window_id) AS max_planned_threads_on_cpu
            FROM solve_windows w ORDER BY w.begin_ns
        """,
        "Windows": "SELECT * FROM solve_windows ORDER BY begin_ns",
        "Threads": "SELECT * FROM threads ORDER BY window_id, demand DESC, tid",
        "Groups": "SELECT * FROM groups ORDER BY window_id, total_demand DESC, thread_count DESC",
        "Top Edges": """
            SELECT * FROM edges
            QUALIFY row_number() OVER (PARTITION BY window_id ORDER BY score DESC, from_tid, to_tid) <= 200
            ORDER BY window_id, score DESC, from_tid, to_tid
        """,
        "Topology CPUs": "SELECT * FROM topology_cpus ORDER BY cpu",
        "Topology Edges": "SELECT * FROM topology_edges ORDER BY from_node, to_node",
        "Plans": "SELECT * FROM plans ORDER BY window_id, strategy_id",
        "Assignments": "SELECT * FROM assignments ORDER BY window_id, strategy_id, cpu, tid",
        "CPU Distribution": "SELECT * FROM cpu_loads ORDER BY window_id, strategy_id, thread_count DESC, cpu",
        "Actions": "SELECT * FROM action_batches ORDER BY timestamp_ns",
        "BPF Health": "SELECT * FROM bpf_health ORDER BY timestamp_ns",
    }
    counts: dict[str, int] = {}
    for sheet_name, query in queries.items():
        cursor = connection.execute(query)
        columns = [item[0] for item in cursor.description]
        sheet = workbook.create_sheet(sheet_name)
        header = []
        for name in columns:
            cell = WriteOnlyCell(sheet, value=name)
            cell.fill = header_fill
            cell.font = header_font
            header.append(cell)
        sheet.append(header)
        count = 0
        while True:
            rows = cursor.fetchmany(5000)
            if not rows:
                break
            for row in rows:
                sheet.append(list(row))
            count += len(rows)
        counts[sheet_name] = count
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{count + 1}"
        for index, name in enumerate(columns, 1):
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(name) + 2, 12), 28)
    connection.close()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return counts
