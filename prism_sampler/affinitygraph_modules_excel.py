from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable, Sequence

import duckdb
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


MODULES = {
    "THREAD LIFECYCLE": {
        "color": "16877C",
        "contract": [
            ("INPUT", "sched_process_fork / exec / exit", "eBPF tracepoint", "available"),
            ("INPUT", "optional pthread_create uprobe", "start_routine address enrichment", "available; optional"),
            ("INPUT", "task_rename", "thread_name events", "available"),
            ("INPUT", "/proc/<pid>/task reconciliation", "pre-existing/missed thread reconciliation", "used at runtime; origin not separable in old export"),
            ("PROCESS", "tracepoints cover pthread/raw clone/clone3/descendants", "Runtime lifecycle map", "implemented"),
            ("PROCESS", "stable identity = (tgid, tid, starttime)", "threads table", "available"),
            ("OUTPUT", "lifecycle events + parent lineage", "runtime.jsonl", "available"),
            ("OUTPUT", "stable thread identity", "threads table", "available"),
            ("OUTPUT", "final comm + start symbol metadata", "runtime.jsonl", "available; start symbol may be empty"),
        ],
    },
    "TRUSTED COLLECTION": {
        "color": "2878B5",
        "contract": [
            ("INPUT", "lifecycle, futex, selective VFS observations", "BPF maps/ring buffer", "available"),
            ("INPUT", "/proc schedstat/state/CPU/context switches/allowed CPUs", "ProcCollector", "partly available in old export"),
            ("INPUT", "CPU/NUMA topology + calibration", "topology tables", "topology available; calibration detail absent in old log"),
            ("PROCESS", "60 s rolling GraphWindow", "thread and edge windows", "available; old 10 s solve framing reconstructed"),
            ("PROCESS", "30 s loss ratio and health gates", "bpf_health", "available"),
            ("OUTPUT", "per-TID demand + relation evidence", "threads / edges", "available; new evidence columns absent"),
            ("OUTPUT", "BPF health + loss telemetry", "bpf_health", "available"),
            ("OUTPUT", "hard failure when core BPF unhealthy", "hard_failure/runtime gate", "no hard failure in this run"),
        ],
    },
    "GRAPH + SOLVER": {
        "color": "C58A17",
        "contract": [
            ("INPUT", "stable TIDs and inferred groups", "threads.group_name", "available"),
            ("INPUT", "d_i = EWMA(run + runqueue)", "threads.demand", "available"),
            ("INPUT", "futex/VFS relationship graph", "edges", "available"),
            ("INPUT", "hardware latency/current placement/migration budget", "topology/current_cpu/config", "available"),
            ("PROCESS", "NUMA partition -> FM -> LPT singleton assignment", "production Solver", "implemented"),
            ("PROCESS", "confidence/dwell/three-window confirmation", "plan confirmation", "available"),
            ("OUTPUT", "per-TID singleton CPU plan", "assignments", "available"),
            ("OUTPUT", "overload/relation/migration objective components", "plans", "available"),
            ("OUTPUT", "migration decision + confidence", "assignment joined with threads", "available"),
        ],
    },
    "SAFE EXECUTION": {
        "color": "36944F",
        "contract": [
            ("INPUT", "confirmed singleton CPU plan", "active runtime planned_assignments", "available"),
            ("INPUT", "live TIDs + startup CPU envelope", "runtime status/topology", "available"),
            ("INPUT", "saved application affinity masks", "Actuator in-memory restore map", "not exported by old log"),
            ("PROCESS", "active-only sched_setaffinity(singleton)", "action batches", "available"),
            ("PROCESS", "verify masks + transactional rollback", "measurement-start masks/action telemetry", "available"),
            ("OUTPUT", "committed active placement", "actual masks", "available"),
            ("OUTPUT", "requested/committed/vanished/rollback", "action events", "available"),
            ("OUTPUT", "verified singleton masks", "hook measurement_start", "available"),
            ("OUTPUT", "restore result required 100%", "pause/runtime_stop", "available: 723/723"),
        ],
    },
}


def _events(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def _safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]", "_", value)


def _write_contract(sheet: Any, module: str, source_note: str) -> int:
    info = MODULES[module]
    sheet.merge_cells("A1:H1")
    sheet["A1"] = module
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=info["color"])
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:H2")
    sheet["A2"] = source_note
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet["A4"] = "MODULE CONTRACT: INPUT -> PROCESS -> OUTPUT"
    sheet["A4"].font = Font(bold=True, color=info["color"])
    headers = ["stage", "item", "source / representation", "canonical-v11 availability"]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(5, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=info["color"])
    for row_index, values in enumerate(info["contract"], 6):
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, value)
    end = 5 + len(info["contract"])
    table = Table(displayName=_safe_name(module) + "_Contract", ref=f"A5:D{end}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    return end + 2


def _write_table(sheet: Any, start: int, title: str, columns: Sequence[str],
                 rows: Iterable[Sequence[Any]], color: str, table_id: int) -> int:
    values = list(rows)
    sheet.cell(start, 1, title).font = Font(bold=True, color=color, size=12)
    header_row = start + 1
    for column, value in enumerate(columns, 1):
        cell = sheet.cell(header_row, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
    for row_index, row in enumerate(values, header_row + 1):
        for column, value in enumerate(row, 1):
            sheet.cell(row_index, column, value)
    end = header_row + max(len(values), 1)
    if not values:
        sheet.cell(header_row + 1, 1, "no rows")
    table = Table(displayName=f"{_safe_name(sheet.title)}_{table_id}",
                  ref=f"A{header_row}:{get_column_letter(len(columns))}{end}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    return end + 2


def _finalize(sheet: Any) -> None:
    sheet.freeze_panes = "A5"
    for column in range(1, sheet.max_column + 1):
        width = 12
        for row in range(1, min(sheet.max_row, 200) + 1):
            value = sheet.cell(row, column).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 42))
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_modules_workbook(dataset: Path, plan_log: Path, active_log: Path,
                            active_result: Path, active_gate: Path,
                            output: Path) -> dict[str, int]:
    connection = duckdb.connect(str(dataset), read_only=True)
    plan_events = _events(plan_log)
    active_events = _events(active_log)
    result = json.loads(active_result.read_text())
    gate = json.loads(active_gate.read_text())
    workbook = Workbook()
    workbook.remove(workbook.active)
    row_counts: dict[str, int] = {}

    # Module 1: authoritative lifecycle input and stable-identity output.
    module = "THREAD LIFECYCLE"
    sheet = workbook.create_sheet(module)
    cursor = _write_contract(sheet, module,
        "Source: canonical-v11 plan runtime.jsonl + fixed placement DuckDB. Old logs do not identify whether a reconciled identity first came from BPF or /proc.")
    counts = Counter(row["type"] for row in plan_events)
    event_summary = [
        ("sched_process_exec", "task_exec", counts["task_exec"], "authoritative tracepoint"),
        ("sched_process_fork", "thread_start", counts["thread_start"], "authoritative tracepoint; optional uprobe enrichment"),
        ("task_rename", "thread_name", counts["thread_name"], "metadata; coalesced by BPF"),
        ("sched_process_exit", "thread_exit", counts["thread_exit"], "authoritative tracepoint"),
        ("/proc reconciliation", "not separately logged", None, "used to recover pre-existing/missed threads"),
    ]
    cursor = _write_table(sheet, cursor, "INPUT - lifecycle source summary",
                          ["kernel/user source", "runtime event", "event_count", "interpretation"],
                          event_summary, MODULES[module]["color"], 1)
    lifecycle: dict[int, dict[str, Any]] = {}
    root_tgid = next(int(row["root_pid"]) for row in plan_events if row["type"] == "runtime_start")
    for event in plan_events:
        tid = event.get("tid")
        if tid is None:
            continue
        item = lifecycle.setdefault(int(tid), {"tgid": root_tgid, "tid": int(tid), "status": "observed"})
        if event["type"] == "thread_start":
            item.update({"tgid": event.get("tgid", root_tgid), "parent_tid": event.get("parent_tid"),
                         "start_routine": event.get("start_routine"), "start_symbol": event.get("start_symbol"),
                         "first_event_ns": event.get("timestamp_ns"), "status": "live"})
        elif event["type"] == "thread_name":
            item["final_comm"] = event.get("name")
        elif event["type"] == "thread_exit":
            item.update({"status": "exited", "last_event_ns": event.get("timestamp_ns")})
    identities = connection.execute("""
        SELECT tgid, tid, min(starttime), max(starttime), count(DISTINCT window_id)
        FROM threads GROUP BY tgid, tid ORDER BY tid
    """).fetchall()
    identity_rows = []
    for tgid, tid, first_start, last_start, windows in identities:
        item = lifecycle.get(tid, {})
        identity_rows.append((tgid, tid, first_start, f"({tgid}, {tid}, {first_start})",
                              item.get("parent_tid"), item.get("final_comm"), item.get("start_routine"),
                              item.get("start_symbol"), item.get("status", "reconciled/observed"), windows,
                              first_start != last_start))
    cursor = _write_table(sheet, cursor, "OUTPUT - stable identities and final metadata",
        ["tgid", "tid", "starttime", "stable_identity", "parent_tid", "final_comm",
         "start_routine", "start_symbol", "final_status", "observed_windows", "tid_reuse_seen"],
        identity_rows, MODULES[module]["color"], 2)
    row_counts[module] = len(identity_rows)
    _finalize(sheet)

    # Module 2: collection inputs, trust telemetry, and selected graph-window outputs.
    module = "TRUSTED COLLECTION"
    sheet = workbook.create_sheet(module)
    cursor = _write_contract(sheet, module,
        "Source: fixed plan DuckDB. Each solve window is legacy_reconstructed; missing old-schema fields remain blank.")
    window_rows = connection.execute("""
        SELECT w.window_id, w.outcome, count(*) AS thread_count,
               count(*) FILTER (WHERE t.confidence >= 0.8) AS eligible_threads,
               sum(t.demand) AS total_demand,
               (SELECT count(*) FROM edges e WHERE e.window_id=w.window_id) AS edge_count
        FROM solve_windows w JOIN threads t USING(window_id)
        GROUP BY w.window_id, w.begin_ns, w.outcome ORDER BY w.begin_ns
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT - reconstructed 60 s graph windows at each 10 s solve",
        ["window_id", "framing", "thread_count", "confidence_ge_0_8", "total_demand", "relation_edges"],
        window_rows, MODULES[module]["color"], 1)
    topology_rows = connection.execute("""
        SELECT 'CPU' kind, cast(cpu as varchar) id, cast(node as varchar) node_or_to,
               cast(online as varchar) metric_value, 'online / envelope' metric
        FROM topology_cpus
        UNION ALL
        SELECT 'NUMA_EDGE', cast(from_node as varchar), cast(to_node as varchar),
               cast(numa_distance as varchar), 'NUMA distance'
        FROM topology_edges ORDER BY kind, id
    """).fetchall()
    cursor = _write_table(sheet, cursor, "INPUT - CPU / NUMA topology",
        ["kind", "cpu_or_from_node", "node_or_to_node", "value", "metric"], topology_rows,
        MODULES[module]["color"], 2)
    health_rows = connection.execute("""
        SELECT timestamp_ns, valid, error, emitted, dropped, loss_ratio, window_ready, final
        FROM bpf_health ORDER BY timestamp_ns
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT - BPF health and loss telemetry",
        ["timestamp_ns", "health_map_valid", "error", "emitted", "dropped", "loss_ratio",
         "30s_window_ready", "final"], health_rows, MODULES[module]["color"], 3)
    demand_rows = connection.execute("""
        SELECT window_id, tgid, tid, starttime, group_name, demand, confidence, current_cpu,
               allowed_cpus, state, sample_count, runtime_delta_ns, runqueue_delta_ns
        FROM threads
        QUALIFY row_number() OVER (PARTITION BY window_id ORDER BY demand DESC, tid) <= 50
        ORDER BY window_id, demand DESC
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT SAMPLE - top 50 demand threads per window",
        ["window_id", "tgid", "tid", "starttime", "group", "demand", "confidence",
         "current_cpu", "allowed_cpus", "state", "sample_count", "runtime_delta_ns", "runqueue_delta_ns"],
        demand_rows, MODULES[module]["color"], 4)
    edge_rows = connection.execute("""
        SELECT window_id, from_tid, from_starttime, to_tid, to_starttime,
               activity, sync, share, stability, score, handoff_rate,
               shared_vfs_seconds, active_overlap, observation_count, coverage, cv
        FROM edges
        QUALIFY row_number() OVER (PARTITION BY window_id ORDER BY score DESC, from_tid, to_tid) <= 50
        ORDER BY window_id, score DESC
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT SAMPLE - top 50 relationship edges per window",
        ["window_id", "from_tid", "from_starttime", "to_tid", "to_starttime", "activity",
         "sync", "share", "stability", "score", "handoff_rate", "shared_vfs_seconds",
         "active_overlap", "observation_count", "coverage", "cv"], edge_rows,
        MODULES[module]["color"], 5)
    row_counts[module] = len(window_rows) + len(health_rows) + len(demand_rows) + len(edge_rows)
    _finalize(sheet)

    # Module 3: solver inputs/outputs and all per-thread migration decisions.
    module = "GRAPH + SOLVER"
    sheet = workbook.create_sheet(module)
    cursor = _write_contract(sheet, module,
        "Source: canonical-v11 plan-mode reconstructed windows. Plans are proposals only; this sheet does not imply affinity execution.")
    plan_rows = connection.execute("""
        WITH distribution AS (
          SELECT window_id, strategy_id, cpu, count(*) AS thread_count, sum(t.demand) AS demand
          FROM assignments a JOIN threads t USING(window_id, tgid, tid, starttime)
          GROUP BY window_id, strategy_id, cpu
        )
        SELECT p.window_id, p.strategy_id, p.thread_count, p.confirmation_threads,
               p.confirmation, p.overload, p.relation_cost, p.migration_cost,
               max(d.thread_count) AS max_threads_per_cpu, max(d.demand) AS max_cpu_demand
        FROM plans p JOIN distribution d USING(window_id, strategy_id)
        GROUP BY ALL ORDER BY p.window_id
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT - plan and objective components",
        ["window_id", "strategy", "planned_threads", "confirmation_threads", "confirmation",
         "overload", "relationship_latency_cost", "migration_cost", "max_threads_per_cpu", "max_cpu_demand"],
        plan_rows, MODULES[module]["color"], 1)
    distribution_rows = connection.execute("""
        SELECT c.window_id, c.strategy_id, c.cpu, t.node, c.thread_count, c.demand
        FROM cpu_loads c LEFT JOIN topology_cpus t USING(cpu)
        ORDER BY c.window_id, c.thread_count DESC, c.cpu
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT - per-CPU plan distribution",
        ["window_id", "strategy", "planned_cpu", "planned_node", "thread_count", "total_demand"],
        distribution_rows, MODULES[module]["color"], 2)
    decision_rows = connection.execute("""
        SELECT a.window_id, a.tgid, a.tid, a.starttime, t.group_name, t.demand, t.confidence,
               t.current_cpu, current_topology.node AS current_node,
               a.cpu AS planned_cpu, planned_topology.node AS planned_node,
               a.cpu != t.current_cpu AS migrate,
               CASE WHEN current_topology.node IS NULL OR planned_topology.node IS NULL THEN NULL
                    ELSE current_topology.node != planned_topology.node END AS cross_node
        FROM assignments a JOIN threads t USING(window_id, tgid, tid, starttime)
        LEFT JOIN topology_cpus current_topology ON current_topology.cpu=t.current_cpu
        LEFT JOIN topology_cpus planned_topology ON planned_topology.cpu=a.cpu
        ORDER BY a.window_id, t.demand DESC, a.tid
    """).fetchall()
    cursor = _write_table(sheet, cursor, "OUTPUT - per-TID singleton plan and migration decision",
        ["window_id", "tgid", "tid", "starttime", "group", "demand", "confidence",
         "current_cpu", "current_node", "planned_singleton_cpu", "planned_node", "migrate", "cross_node"],
        decision_rows, MODULES[module]["color"], 3)
    row_counts[module] = len(plan_rows) + len(distribution_rows) + len(decision_rows)
    _finalize(sheet)

    # Module 4: real active evidence, kept separate from plan-mode facts.
    module = "SAFE EXECUTION"
    sheet = workbook.create_sheet(module)
    cursor = _write_contract(sheet, module,
        "Source: canonical-v11 active runtime/result/gate. Environment was busy, so action safety evidence is usable but throughput is not an algorithm label.")
    gate_rows = [(key, value) for key, value in gate.items() if not isinstance(value, (dict, list))]
    gate_rows.extend([
        ("environment_valid", False),
        ("throughput_label_usable", False),
        ("safety_interpretation", "action/restore evidence only; busy host blocks performance conclusion"),
    ])
    cursor = _write_table(sheet, cursor, "OUTPUT - active safety gate summary",
        ["metric", "value"], gate_rows, MODULES[module]["color"], 1)
    action_rows = []
    restore_rows = []
    inherit_rows = []
    for event in active_events:
        if event["type"] == "action":
            action_rows.append((event.get("timestamp_ns"), len(event.get("assignments", {})),
                                event.get("requested"), event.get("applied"), event.get("committed"),
                                event.get("vanished"), event.get("rolled_back"),
                                event.get("rollback_success"), event.get("error"), event.get("success")))
        elif event["type"] in {"pause", "runtime_stop"}:
            restore_rows.append((event.get("timestamp_ns"), event["type"], event.get("restore_requested"),
                                 event.get("restore_restored"), event.get("restore_vanished"),
                                 event.get("restore_failed"), event.get("restored")))
        elif event["type"] == "thread_inherit_plan":
            inherit_rows.append((event.get("timestamp_ns"), event.get("tid"), event.get("cpu"),
                                 event.get("success"), event.get("error")))
    cursor = _write_table(sheet, cursor, "OUTPUT - transactional action batches",
        ["timestamp_ns", "assignment_count", "requested", "applied", "committed", "vanished",
         "rolled_back", "rollback_success", "error", "success"], action_rows,
        MODULES[module]["color"], 2)
    cursor = _write_table(sheet, cursor, "OUTPUT - new-thread inherited plan actions",
        ["timestamp_ns", "tid", "cpu", "success", "error"], inherit_rows,
        MODULES[module]["color"], 3)
    measurement = result["measurement_start"]
    actual_masks = measurement.get("masks", {})
    planned = measurement.get("runtime_status", {}).get("planned_assignments", {})
    mask_rows = []
    for tid_text, mask in sorted(actual_masks.items(), key=lambda item: int(item[0])):
        cpus = [part for part in re.split(r"[,-]", str(mask)) if part]
        planned_cpu = planned.get(tid_text, planned.get(str(tid_text)))
        actual_cpu = int(cpus[0]) if len(cpus) == 1 and cpus[0].isdigit() else None
        mask_rows.append((int(tid_text), str(mask), len(cpus) == 1, planned_cpu, actual_cpu,
                          actual_cpu == planned_cpu if actual_cpu is not None and planned_cpu is not None else None))
    cursor = _write_table(sheet, cursor, "OUTPUT - measurement-start verified singleton masks",
        ["tid", "actual_affinity_mask", "singleton_verified", "planned_cpu", "actual_cpu", "matches_plan"],
        mask_rows, MODULES[module]["color"], 4)
    cursor = _write_table(sheet, cursor, "OUTPUT - restore batches",
        ["timestamp_ns", "event", "restore_requested", "restore_restored", "restore_vanished",
         "restore_failed", "restored"], restore_rows, MODULES[module]["color"], 5)
    row_counts[module] = len(action_rows) + len(inherit_rows) + len(mask_rows) + len(restore_rows)
    _finalize(sheet)

    connection.close()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    # A second parser pass catches malformed workbook packages before handoff.
    checked = load_workbook(output, read_only=True, data_only=True)
    if checked.sheetnames != list(MODULES):
        raise RuntimeError(f"unexpected module sheets: {checked.sheetnames}")
    checked.close()
    return row_counts
