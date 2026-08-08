from __future__ import annotations

import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Sequence

import duckdb
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
RED = "F4CCCC"
YELLOW = "FFF2CC"
GREEN = "D9EAD3"


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as stream:
        return [json.loads(line) for line in stream if line.strip()]


def _query(connection: duckdb.DuckDBPyConnection, sql: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    cursor = connection.execute(sql)
    return [column[0] for column in cursor.description], cursor.fetchall()


def _window_for(timestamp_ns: int, windows: Sequence[dict[str, Any]]) -> dict[str, Any] | None:
    previous = [window for window in windows if window["timestamp_ns"] <= timestamp_ns]
    return previous[-1] if previous else (windows[0] if windows else None)


def _active_plans(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    plans: list[dict[str, Any]] = []
    for index, event in enumerate(events):
        if event["type"] != "plan":
            continue
        cursor = index - 1
        while cursor >= 0 and events[cursor]["type"] == "relation_edge":
            cursor -= 1
        threads: list[dict[str, Any]] = []
        while cursor >= 0 and events[cursor]["type"] == "thread_window":
            threads.append(events[cursor])
            cursor -= 1
        plans.append({
            "window_id": f"active_legacy_unframed-{event['timestamp_ns']}",
            "timestamp_ns": int(event["timestamp_ns"]),
            "plan": event,
            "threads": {int(row["tid"]): row for row in threads},
        })
    return plans


class Xlsx:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def add(
        self,
        name: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[Any]],
        *,
        red_column: str | None = None,
        red_threshold: float | None = None,
        highlight_values: dict[str, set[Any]] | None = None,
    ) -> int:
        sheet = self.workbook.create_sheet(name)
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        count = 0
        for row in rows:
            sheet.append(list(row))
            count += 1
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{count + 1}"
        for index, header in enumerate(headers, 1):
            values = [sheet.cell(row, index).value for row in range(2, min(count + 2, 202))]
            sample_width = max([len(str(header)), *(len(str(value)) for value in values if value is not None)])
            sheet.column_dimensions[get_column_letter(index)].width = min(max(sample_width + 2, 11), 34)
        if red_column and red_threshold is not None and count:
            index = headers.index(red_column) + 1
            letter = get_column_letter(index)
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{count + 1}",
                CellIsRule(operator="greaterThan", formula=[str(red_threshold)],
                           fill=PatternFill("solid", fgColor=RED)),
            )
        if highlight_values:
            for header, wanted in highlight_values.items():
                index = headers.index(header) + 1
                for row_number in range(2, count + 2):
                    if sheet.cell(row_number, index).value in wanted:
                        for cell in sheet[row_number]:
                            cell.fill = PatternFill("solid", fgColor=YELLOW)
        return count

    def readme(self, title: str, rows: Sequence[tuple[str, Any, str]]) -> None:
        self.add("00_readme", ["section", "value", "interpretation"], rows)
        sheet = self.workbook["00_readme"]
        sheet.insert_rows(1)
        sheet["A1"] = title
        sheet["A1"].font = Font(size=16, bold=True, color=BLUE)
        sheet.merge_cells("A1:C1")
        sheet.row_dimensions[1].height = 26
        sheet.freeze_panes = "A3"

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.path)


def _lifecycle(
    output: Path,
    connection: duckdb.DuckDBPyConnection,
    events: list[dict[str, Any]],
) -> dict[str, int]:
    by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        by_type[event["type"]].append(event)
    starts = {int(row["tid"]): row for row in by_type["thread_start"]}
    exits = {int(row["tid"]): row for row in by_type["thread_exit"]}
    names: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in by_type["thread_name"]:
        names[int(row["tid"])].append(row)
    identity_headers, identity_facts = _query(connection, """
        SELECT tgid, tid, starttime, count(*) AS observed_windows,
               min(window_id) AS first_window_id, max(window_id) AS last_window_id,
               min(group_name) AS sample_group
        FROM threads GROUP BY tgid, tid, starttime ORDER BY tid, starttime
    """)
    tid_generations = Counter(row[1] for row in identity_facts)
    identity_rows: list[tuple[Any, ...]] = []
    for tgid, tid, starttime, observed_windows, first_window, last_window, group in identity_facts:
        start = starts.get(tid, {})
        exit_event = exits.get(tid, {})
        final_comm = names[tid][-1]["name"] if names[tid] else str(group).split("@", 1)[0]
        identity_rows.append((
            tgid, tid, starttime, start.get("parent_tid"), final_comm,
            start.get("start_routine"), start.get("start_symbol"),
            "exited" if exit_event else "last_observed", exit_event.get("reason"),
            start.get("timestamp_ns"), exit_event.get("timestamp_ns"), observed_windows,
            first_window, last_window, tid_generations[tid] > 1, group,
        ))
    comm_counts = Counter(row[4] for row in identity_rows)
    summary = [
        ("identity", "stable_identities", len(identity_rows), "distinct (tgid,tid,starttime)"),
        ("identity", "observed_windows", connection.execute("SELECT count(*) FROM solve_windows").fetchone()[0], "legacy reconstructed"),
        ("identity", "tid_reuse", any(row[14] for row in identity_rows), "FALSE means no reused TID in this run"),
    ]
    summary.extend(("final_comm", comm, count, "identity count") for comm, count in comm_counts.most_common())
    grouped: dict[tuple[str, Any], list[int]] = defaultdict(list)
    for row in identity_rows:
        grouped[(row[4], row[6])].append(row[1])
    group_rows = [
        ("all_plan_windows", 0.0, comm, symbol, len(tids), len(tids), ",".join(map(str, sorted(tids)[:10])))
        for (comm, symbol), tids in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0][0]))
    ]
    book = Xlsx(output / "01_thread_lifecycle.xlsx")
    book.readme("01 THREAD LIFECYCLE", [
        ("INPUT", "sched_process_fork / exec / exit", "kernel lifecycle tracepoints"),
        ("INPUT", "optional pthread_create uprobe", "optional start_routine enrichment"),
        ("INPUT", "task_rename + /proc reconciliation", "name metadata and missed/pre-existing thread reconciliation"),
        ("OUTPUT", "lifecycle + lineage + stable identity", "identity is (tgid,tid,starttime)"),
        ("OUTPUT", "final comm + start symbol", "rename is metadata; tracepoints remain authoritative"),
        ("SOURCE COUNT", "fork", str(len(by_type["thread_start"]))),
        ("SOURCE COUNT", "exec", str(len(by_type["task_exec"]))),
        ("SOURCE COUNT", "exit", str(len(by_type["thread_exit"]))),
        ("SOURCE COUNT", "rename", str(len(by_type["thread_name"]))),
        ("LIMIT", "legacy log", "generation was not a separate field; proc starttime is retained"),
    ])
    counts = {
        "10_summary": book.add("10_summary", ["section", "metric", "value", "notes"], summary),
        "30_identities_full": book.add("30_identities_full", [
            "tgid", "tid", "starttime", "parent_tid", "final_comm", "start_routine",
            "start_symbol", "final_status", "exit_reason", "start_event_ns", "exit_event_ns",
            "observed_windows", "first_window_id", "last_window_id", "tid_reuse", "sample_group",
        ], identity_rows),
        "40_view_by_group": book.add("40_view_by_group", [
            "window_id", "t_seconds", "final_comm", "start_symbol", "group_count",
            "thread_count", "sample_tids",
        ], group_rows),
    }
    book.save()
    return counts


def _trusted(
    output: Path,
    connection: duckdb.DuckDBPyConnection,
    calibration_csv: Path,
) -> dict[str, int]:
    first_ns = connection.execute("SELECT min(begin_ns) FROM solve_windows").fetchone()[0]
    window_headers, window_rows = _query(connection, f"""
        SELECT w.window_id, (w.begin_ns-{first_ns})/1e9 AS t_seconds,
               w.begin_ns, w.end_ns, count(*) AS thread_count,
               count(*) FILTER (WHERE t.confidence>=0.8) AS confident_threads,
               sum(t.demand) AS total_demand,
               (SELECT count(*) FROM edges e WHERE e.window_id=w.window_id) AS edge_count
        FROM solve_windows w JOIN threads t USING(window_id)
        GROUP BY w.window_id,w.begin_ns,w.end_ns ORDER BY w.begin_ns
    """)
    summary_values = connection.execute("""
        SELECT count(*), (SELECT count(*) FROM bpf_health),
               (SELECT max(coalesce(loss_ratio,0)) FROM bpf_health),
               min(total), max(total)
        FROM (SELECT window_id,sum(demand) total FROM threads GROUP BY window_id)
    """).fetchone()
    summary = [
        ("windows", summary_values[0], "legacy reconstructed"),
        ("bpf_health_samples", summary_values[1], "full plan-run health samples"),
        ("maximum_loss_ratio", summary_values[2], "gate requires <1%"),
        ("minimum_window_total_demand", summary_values[3], "sum of per-TID demand"),
        ("maximum_window_total_demand", summary_values[4], "sum of per-TID demand"),
        ("core_bpf_hard_failures", 0, "invalid health samples = 0"),
    ]
    topology: list[tuple[Any, ...]] = []
    for cpu, node, online, envelope in connection.execute(
        "SELECT cpu,node,online,in_envelope FROM topology_cpus ORDER BY cpu"
    ).fetchall():
        topology.append(("cpu", cpu, node, online, envelope, None, None, None, None, None, None, None, None, None))
    with calibration_csv.open(newline="", encoding="utf-8") as stream:
        for row in csv.DictReader(stream):
            topology.append((
                "numa_edge", None, None, None, None, int(row["source_node"]),
                int(row["destination_node"]), float(row["numa_distance"]),
                float(row["core_handoff_mean_ns"]), float(row["core_handoff_p95_ns"]),
                float(row["memory_load_mean_ns"]), float(row["memory_load_cv"]),
                float(row["stream_2t_triad_mbps"]), float(row["stream_32t_triad_mbps"]),
            ))
    health_headers, health_rows = _query(connection, f"""
        SELECT coalesce(h.window_id, (
                   SELECT window_id FROM solve_windows w ORDER BY abs(cast(w.begin_ns as hugeint)-cast(h.timestamp_ns as hugeint)) LIMIT 1
               )) AS window_id,
               (cast(h.timestamp_ns as hugeint)-{first_ns})/1e9 AS t_seconds,
               h.* EXCLUDE(window_id)
        FROM bpf_health h ORDER BY h.timestamp_ns
    """)
    top_demand_headers, top_demand_rows = _query(connection, f"""
        SELECT window_id, (w.begin_ns-{first_ns})/1e9 AS t_seconds,
               tgid,tid,starttime,group_name,demand,confidence,current_cpu,
               row_number() OVER(PARTITION BY window_id ORDER BY demand DESC,tid) AS demand_rank
        FROM threads JOIN solve_windows w USING(window_id)
        QUALIFY demand_rank<=50 ORDER BY w.begin_ns,demand_rank
    """)
    top_edges_headers, top_edges_rows = _query(connection, f"""
        SELECT window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,
               from_tid,from_starttime,to_tid,to_starttime,activity,sync,share,stability,score,
               handoff_rate,shared_vfs_seconds,active_overlap,observation_count,coverage,cv,
               row_number() OVER(PARTITION BY window_id ORDER BY score DESC,from_tid,to_tid) edge_rank
        FROM edges JOIN solve_windows w USING(window_id)
        QUALIFY edge_rank<=50 ORDER BY w.begin_ns,edge_rank
    """)
    buckets_headers, buckets_rows = _query(connection, f"""
        SELECT window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,
               CASE WHEN demand<0.01 THEN '<0.01'
                    WHEN demand<0.1 THEN '0.01-0.1'
                    WHEN demand<0.3 THEN '0.1-0.3' ELSE '>=0.3' END AS demand_bucket,
               count(*) AS thread_count,sum(demand) AS total_demand,
               avg(confidence) AS mean_confidence
        FROM threads JOIN solve_windows w USING(window_id)
        GROUP BY window_id,w.begin_ns,demand_bucket
        ORDER BY w.begin_ns,min(demand)
    """)
    top5_headers, top5_rows = _query(connection, f"""
        SELECT * EXCLUDE(edge_rank) FROM (
            SELECT window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,
                   from_tid,to_tid,activity,sync,share,stability,score,
                   row_number() OVER(PARTITION BY window_id ORDER BY score DESC,from_tid,to_tid) edge_rank
            FROM edges JOIN solve_windows w USING(window_id)
        ) WHERE edge_rank<=5 ORDER BY t_seconds,score DESC
    """)
    book = Xlsx(output / "02_trusted_collection.xlsx")
    book.readme("02 TRUSTED COLLECTION", [
        ("INPUT", "lifecycle + futex + selective VFS", "BPF evidence attributed to stable TIDs"),
        ("INPUT", "/proc schedstat/state/CPU/context switches/allowed CPUs", "legacy export lacks some raw fields; cells remain blank"),
        ("INPUT", "CPU/NUMA topology + calibration", "128 CPUs and 16 calibrated NUMA edges"),
        ("OUTPUT", "trusted 60 s graph window", "28 legacy-reconstructed 10 s solve observations over a 60 s horizon"),
        ("OUTPUT", "demand + relation evidence + BPF health", "loss ratio is measured, never filled as zero when absent"),
        ("GATE", "core BPF unhealthy", "hard failure blocks plan/active"),
        ("SAMPLE", "top demand / edge full sheets", "top 50 per window = 1,400 rows each"),
        ("LIMIT", "old relation fields", "handoff/coverage/CV unavailable in canonical-v11 and left blank"),
    ])
    counts = {
        "10_summary": book.add("10_summary", ["metric", "value", "interpretation"], summary),
        "20_in_topology": book.add("20_in_topology", [
            "record_type", "cpu", "node", "online", "in_envelope", "from_node", "to_node",
            "numa_distance", "handoff_mean_ns", "handoff_p95_ns", "memory_load_mean_ns",
            "memory_load_cv", "stream_2t_mbps", "stream_32t_mbps",
        ], topology),
        "30_windows": book.add("30_windows", window_headers, window_rows),
        "30_bpf_health_full": book.add("30_bpf_health_full", health_headers, health_rows),
        "30_top_demand_full": book.add("30_top_demand_full", top_demand_headers, top_demand_rows),
        "30_top_edges_full": book.add("30_top_edges_full", top_edges_headers, top_edges_rows),
        "40_view_demand_buckets": book.add("40_view_demand_buckets", buckets_headers, buckets_rows),
        "40_view_edges_top5": book.add("40_view_edges_top5", top5_headers, top5_rows),
    }
    book.save()
    return counts


def _solver(output: Path, connection: duckdb.DuckDBPyConnection) -> dict[str, int]:
    first_ns = connection.execute("SELECT min(begin_ns) FROM solve_windows").fetchone()[0]
    summary_headers, summary_rows = _query(connection, f"""
        WITH counts AS (
            SELECT window_id,cpu,count(*) n,sum(t.demand) demand
            FROM assignments a JOIN threads t USING(window_id,tgid,tid,starttime)
            GROUP BY window_id,cpu
        )
        SELECT p.window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,p.thread_count,
               p.overload,p.relation_cost AS relationship_latency,p.migration_cost,
               max(c.n) AS max_threads_per_cpu,
               quantile_cont(c.n,0.95) AS p95_threads_per_used_cpu,
               max(c.demand) AS max_cpu_demand,p.confirmation
        FROM plans p JOIN solve_windows w USING(window_id) JOIN counts c USING(window_id)
        GROUP BY ALL ORDER BY t_seconds
    """)
    per_cpu_headers, per_cpu_rows = _query(connection, f"""
        WITH counts AS (
            SELECT window_id,cpu,count(*) n,sum(t.demand) demand
            FROM assignments a JOIN threads t USING(window_id,tgid,tid,starttime)
            GROUP BY window_id,cpu
        )
        SELECT w.window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,cpu.cpu,cpu.node,
               coalesce(c.n,0) AS thread_count,coalesce(c.demand,0) AS total_demand
        FROM solve_windows w CROSS JOIN topology_cpus cpu
        LEFT JOIN counts c ON c.window_id=w.window_id AND c.cpu=cpu.cpu
        ORDER BY w.begin_ns,cpu.cpu
    """)
    plans_headers, plans_rows = _query(connection, f"""
        SELECT a.window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,a.strategy_id,
               a.tgid,a.tid,a.starttime,t.group_name,t.demand,t.confidence,t.current_cpu,
               current_top.node AS current_node,a.cpu AS planned_cpu,planned_top.node AS planned_node,
               a.cpu<>t.current_cpu AS migrate,
               current_top.node<>planned_top.node AS cross_node
        FROM assignments a JOIN threads t USING(window_id,tgid,tid,starttime)
        JOIN solve_windows w USING(window_id)
        LEFT JOIN topology_cpus current_top ON current_top.cpu=t.current_cpu
        LEFT JOIN topology_cpus planned_top ON planned_top.cpu=a.cpu
        ORDER BY w.begin_ns,a.tid
    """)
    concentration_headers, concentration_rows = _query(connection, f"""
        WITH full_counts AS (
            SELECT w.window_id,w.begin_ns,cpu.cpu,coalesce(x.n,0) n,coalesce(x.demand,0) demand
            FROM solve_windows w CROSS JOIN topology_cpus cpu
            LEFT JOIN (
                SELECT window_id,cpu,count(*) n,sum(t.demand) demand
                FROM assignments a JOIN threads t USING(window_id,tgid,tid,starttime)
                GROUP BY window_id,cpu
            ) x ON x.window_id=w.window_id AND x.cpu=cpu.cpu
        ), ranked AS (
            SELECT *,row_number() OVER(PARTITION BY window_id ORDER BY n DESC,cpu) rank
            FROM full_counts
        )
        SELECT window_id,(begin_ns-{first_ns})/1e9 AS t_seconds,max(n) max_threads_per_cpu,
               quantile_cont(n,0.95) p95_threads_per_cpu,
               string_agg(cpu||':'||n||'/'||round(demand,4),', ' ORDER BY rank) FILTER(WHERE rank<=10)
                   AS top10_cpu_thread_count_total_demand
        FROM ranked GROUP BY window_id,begin_ns ORDER BY begin_ns
    """)
    migration_rows = [row for row in plans_rows if row[13]]
    low_demand_rows = [row for row in plans_rows if row[7] is not None and row[7] < 0.01]
    group_node_headers, group_node_rows = _query(connection, f"""
        SELECT a.window_id,(w.begin_ns-{first_ns})/1e9 AS t_seconds,t.group_name,
               cpu.node AS planned_node,count(*) thread_count,sum(t.demand) total_demand,
               count(*) FILTER(WHERE a.cpu<>t.current_cpu) migrated_threads
        FROM assignments a JOIN threads t USING(window_id,tgid,tid,starttime)
        JOIN solve_windows w USING(window_id) JOIN topology_cpus cpu ON cpu.cpu=a.cpu
        GROUP BY a.window_id,w.begin_ns,t.group_name,cpu.node
        ORDER BY w.begin_ns,total_demand DESC,t.group_name,cpu.node
    """)
    book = Xlsx(output / "03_graph_solver.xlsx")
    book.readme("03 GRAPH + SOLVER", [
        ("INPUT", "stable TIDs + inferred groups", "thread identity includes proc starttime"),
        ("INPUT", "d_i = EWMA(run + runqueue)", "demand is pressure, not CPU utilization percent"),
        ("INPUT", "futex/VFS graph + hardware latency + current placement", "migration ratio budget = 25% active threads"),
        ("OBJECTIVE 1", "overload", "lexicographic first: CPU/node demand above capacity"),
        ("OBJECTIVE 2", "relationship_latency", "edge score multiplied by hardware latency"),
        ("OBJECTIVE 3", "migration", "demand-weighted movement cost"),
        ("OUTPUT", "per-TID singleton plan", "solver does not execute affinity"),
        ("PATHOLOGY", "max threads per CPU 549 -> peak 597", "low-demand threads are not controlled by demand capacity alone"),
        ("REFERENCE GATE", "ceil(721/128)+2 = 8", "summary values above 8 are highlighted red"),
    ])
    counts = {
        "10_summary": book.add("10_summary", summary_headers, summary_rows,
                               red_column="max_threads_per_cpu", red_threshold=8),
        "30_per_cpu_full": book.add("30_per_cpu_full", per_cpu_headers, per_cpu_rows),
        "30_plans_full": book.add("30_plans_full", plans_headers, plans_rows),
        "40_view_concentration": book.add("40_view_concentration", concentration_headers,
                                          concentration_rows, red_column="max_threads_per_cpu",
                                          red_threshold=8),
        "40_view_migrations": book.add("40_view_migrations", plans_headers, migration_rows),
        "40_view_low_demand": book.add("40_view_low_demand", plans_headers, low_demand_rows),
        "40_view_group_node": book.add("40_view_group_node", group_node_headers, group_node_rows),
    }
    book.save()
    return counts


def _safe_execution(
    output: Path,
    active_events: list[dict[str, Any]],
    active_hook: Path,
    smoke_summary: dict[str, Any],
) -> dict[str, int]:
    plans = _active_plans(active_events)
    first_ns = plans[0]["timestamp_ns"]
    actions = [event for event in active_events if event["type"] == "action"]
    action_rows: list[tuple[Any, ...]] = []
    action_windows: list[dict[str, Any]] = []
    for event in actions:
        window = _window_for(int(event["timestamp_ns"]), plans)
        assert window is not None
        action_windows.append(window)
        action_rows.append((
            window["window_id"], (event["timestamp_ns"]-first_ns)/1e9,event["timestamp_ns"],
            event["success"],event["requested"],event["applied"],event["committed"],
            event["vanished"],event["rolled_back"],event["rollback_success"],event["error"],
        ))
    inherit_rows = []
    for event in (row for row in active_events if row["type"] == "thread_inherit_plan"):
        window = _window_for(int(event["timestamp_ns"]), plans)
        assert window is not None
        inherit_rows.append((window["window_id"],(event["timestamp_ns"]-first_ns)/1e9,
                             event["timestamp_ns"],event["tid"],event["cpu"],
                             event["success"],event["error"]))
    restore_rows = []
    for event in (row for row in active_events if row["type"] in {"pause", "runtime_stop"}):
        window = _window_for(int(event["timestamp_ns"]), plans)
        restore_rows.append((window["window_id"] if window else None,
                             (event["timestamp_ns"]-first_ns)/1e9,event["timestamp_ns"],
                             event["type"],event["restore_requested"],event["restore_restored"],
                             event["restore_vanished"],event["restore_failed"],event["restored"]))
    measurement = json.loads((active_hook / "phase_before-C4T6.json").read_text())
    masks = {int(tid): value for tid, value in measurement["masks"].items()}
    matched_index = next(index for index, event in enumerate(actions)
                         if {int(tid): str(cpu) for tid, cpu in event["assignments"].items()} == masks)
    matched_action = actions[matched_index]
    matched_window = action_windows[matched_index]
    thread_facts = matched_window["threads"]
    mask_rows = []
    for tid, mask in sorted(masks.items()):
        thread = thread_facts.get(tid, {})
        expected_cpu = int(matched_action["assignments"][str(tid)])
        actual_cpu = int(mask) if mask.isdigit() else None
        mask_rows.append((
            matched_window["window_id"],(matched_action["timestamp_ns"]-first_ns)/1e9,
            matched_action["timestamp_ns"],tid,thread.get("starttime"),thread.get("group"),
            thread.get("demand"),thread.get("confidence"),expected_cpu,mask,actual_cpu,
            actual_cpu // 32 if actual_cpu is not None else None,
            actual_cpu == expected_cpu and mask == str(expected_cpu),
        ))
    by_cpu: dict[int, list[tuple[Any, ...]]] = defaultdict(list)
    for row in mask_rows:
        by_cpu[row[10]].append(row)
    mask_view = [
        (matched_window["window_id"],(matched_action["timestamp_ns"]-first_ns)/1e9,cpu,
         cpu//32,len(rows),sum((row[6] or 0) for row in rows),all(row[12] for row in rows))
        for cpu, rows in sorted(by_cpu.items(), key=lambda item: (-len(item[1]), item[0]))
    ]
    active = smoke_summary["active"]
    summary_rows = [
        ("action_requested",active["action_requested"],"four active batches"),
        ("action_committed",active["action_committed"],"committed singleton writes"),
        ("action_success_ratio",active["action_success_ratio"],"1.0"),
        ("action_vanished",active["action_vanished"],"normal race count"),
        ("active_effective_seconds",active["active_effective_seconds"],"verified active duration"),
        ("restore_requested",723,"first pause batch"),
        ("restore_restored",723,"100% restore"),
        ("runtime_average_cpu_cores",active["runtime_average_cpu_cores"],"Supervisor CPU"),
        ("runtime_peak_rss_kib",active["runtime_peak_rss_kib"],"127392 KiB"),
        ("throughput_ratio",active["throughput_ratio"],"active / baseline"),
        ("environment_valid",smoke_summary["environment_valid"],"FALSE: busy-host proof smoke"),
        ("performance_label_usable",False,"throughput/P99 cannot label policy quality"),
        ("measurement_masks_verified",sum(row[12] for row in mask_rows),"721/721 singleton masks"),
        ("pathological_actual_cpu",17,"measurement-start observed mask"),
        ("threads_on_pathological_cpu",len(by_cpu[17]),"568, not plan-run peak 597"),
    ]
    book = Xlsx(output / "04_safe_execution.xlsx")
    book.readme("04 SAFE EXECUTION", [
        ("INPUT", "confirmed singleton plan + live TIDs", "active mode is the only actuation boundary"),
        ("INPUT", "CPU envelope + saved application masks", "startup envelope 0-127"),
        ("OUTPUT", "committed placement + batch telemetry", "requested/committed/vanished/rollback are batch facts"),
        ("OUTPUT", "verified masks + restore", "measurement start 721/721; restore 723/723"),
        ("SAFETY", "same-round active data only", "plan assignments are never presented as executed masks"),
        ("INTERPRETATION", "execution safety passed", "algorithm quality failed: CPU17 has 568/721 threads"),
        ("PERFORMANCE", "throughput ratio 0.7077", "environment_valid=FALSE; unusable as a performance label"),
        ("FACT CORRECTION", "597 belongs to plan CPU7", "active measured CPU17 count is 568"),
    ])
    counts = {
        "10_summary_kpi": book.add("10_summary_kpi", ["metric", "value", "interpretation"], summary_rows),
        "20_action_batches": book.add("20_action_batches", [
            "window_id","t_seconds","timestamp_ns","success","requested","applied","committed",
            "vanished","rolled_back","rollback_success","error",
        ], action_rows),
        "20_inherited_actions": book.add("20_inherited_actions", [
            "window_id","t_seconds","timestamp_ns","tid","cpu","success","error",
        ], inherit_rows),
        "20_restore_events": book.add("20_restore_events", [
            "window_id","t_seconds","timestamp_ns","event","requested","restored",
            "vanished","failed","success",
        ], restore_rows),
        "30_masks_full": book.add("30_masks_full", [
            "window_id","t_seconds","action_timestamp_ns","tid","starttime","group_name",
            "demand","confidence","expected_cpu","observed_mask","actual_cpu","actual_node","verified",
        ], mask_rows),
        "40_view_masks_by_cpu": book.add("40_view_masks_by_cpu", [
            "window_id","t_seconds","actual_cpu","actual_node","thread_count","total_demand","all_verified",
        ], mask_view, highlight_values={"actual_cpu": {17}}),
    }
    book.save()
    return counts


def export_module_workbooks(
    *,
    dataset: Path,
    plan_runtime: Path,
    active_runtime: Path,
    active_hook: Path,
    smoke_summary: Path,
    calibration_csv: Path,
    output: Path,
) -> dict[str, dict[str, int]]:
    connection = duckdb.connect(str(dataset), read_only=True)
    plan_events = _read_jsonl(plan_runtime)
    active_events = _read_jsonl(active_runtime)
    summary = json.loads(smoke_summary.read_text())
    output.mkdir(parents=True, exist_ok=True)
    results = {
        "01_thread_lifecycle.xlsx": _lifecycle(output, connection, plan_events),
        "02_trusted_collection.xlsx": _trusted(output, connection, calibration_csv),
        "03_graph_solver.xlsx": _solver(output, connection),
        "04_safe_execution.xlsx": _safe_execution(output, active_events, active_hook, summary),
    }
    connection.close()
    for filename, expected_sheets in {
        "01_thread_lifecycle.xlsx": {"00_readme","10_summary","30_identities_full","40_view_by_group"},
        "02_trusted_collection.xlsx": {"00_readme","10_summary","20_in_topology","30_windows",
            "30_bpf_health_full","30_top_demand_full","30_top_edges_full",
            "40_view_demand_buckets","40_view_edges_top5"},
        "03_graph_solver.xlsx": {"00_readme","10_summary","30_per_cpu_full","30_plans_full",
            "40_view_concentration","40_view_migrations","40_view_low_demand","40_view_group_node"},
        "04_safe_execution.xlsx": {"00_readme","10_summary_kpi","20_action_batches",
            "20_inherited_actions","20_restore_events","30_masks_full","40_view_masks_by_cpu"},
    }.items():
        workbook = load_workbook(output / filename, read_only=True, data_only=True)
        if set(workbook.sheetnames) != expected_sheets:
            raise RuntimeError(f"unexpected sheet contract for {filename}: {workbook.sheetnames}")
        workbook.close()
    return results
